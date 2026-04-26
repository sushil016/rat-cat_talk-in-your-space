from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2E4057")
ALT_FILL = PatternFill("solid", fgColor="F4F7FA")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)


def _style_table(ws, header_row: int, data_start_row: int, data_end_row: int, total_cols: int) -> None:
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in range(data_start_row, data_end_row + 1):
        if row % 2 == 0:
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).fill = ALT_FILL


def generate_xlsx_report(df: pd.DataFrame, stats: Dict[str, object], output_dir: str) -> str:
    output_path = str(Path(output_dir) / "cav_report.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("Summary")
    summary_data = [
        ("Total Customers", stats.get("total_customers", 0)),
        ("Total Revenue", float(stats.get("total_revenue", 0.0))),
        ("Average LTV", float(stats.get("avg_ltv", 0.0))),
        ("Top Channel", stats.get("top_channel", "N/A")),
        ("Avg Revenue per Customer", float(stats.get("avg_revenue_per_customer", 0.0))),
        ("Tracked Months", len(stats.get("monthly_acquisitions", {}))),
    ]
    ws.append(["Metric", "Value"])
    for row in summary_data:
        ws.append(list(row))
    _style_table(ws, 1, 2, len(summary_data) + 1, 2)
    _auto_width(ws)

    # By Channel
    ws = wb.create_sheet("By Channel")
    ws.append(["Channel", "Revenue", "Avg LTV", "Customers"])
    channels = set(stats.get("revenue_by_channel", {}).keys()) | set(
        stats.get("ltv_by_channel", {}).keys()
    )
    for channel in sorted(channels):
        ws.append(
            [
                channel,
                float(stats.get("revenue_by_channel", {}).get(channel, 0.0)),
                float(stats.get("ltv_by_channel", {}).get(channel, 0.0)),
                int(stats.get("customer_count_by_channel", {}).get(channel, 0)),
            ]
        )
    _style_table(ws, 1, 2, ws.max_row, 4)
    _auto_width(ws)

    # Monthly Trend
    ws = wb.create_sheet("Monthly Trend")
    ws.append(["Month", "Acquisitions"])
    for month, count in stats.get("monthly_acquisitions", {}).items():
        ws.append([month, int(count)])
    _style_table(ws, 1, 2, ws.max_row, 2)
    _auto_width(ws)

    # Top Customers
    ws = wb.create_sheet("Top Customers")
    ws.append(["Customer ID", "Name", "Channel", "Revenue", "LTV", "Month"])
    for row in stats.get("top_10_customers", []):
        ws.append(
            [
                row.get("customer_id", ""),
                row.get("name", ""),
                row.get("channel", ""),
                float(row.get("revenue", 0.0)),
                float(row.get("ltv", 0.0)),
                row.get("month_year", "Unknown"),
            ]
        )
    _style_table(ws, 1, 2, ws.max_row, 6)
    _auto_width(ws)

    # Raw Normalized
    ws = wb.create_sheet("Raw Normalized")
    ws.append(df.columns.tolist())
    for row in df.fillna("").itertuples(index=False):
        ws.append(list(row))
    _style_table(ws, 1, 2, ws.max_row, len(df.columns))
    _auto_width(ws)

    wb.save(output_path)
    return output_path
